
#This requires python 3.
#On OSX, run export OBJC_DISABLE_INITIALIZE_FORK_SAFETY=YES first
import requests
from google.cloud import spanner #Spanner
from requests.auth import HTTPBasicAuth
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from urllib.parse import unquote
from urllib.parse import urlparse
from urllib.parse import parse_qs
from multiprocessing import Process
from datetime import datetime, timezone, timedelta
import fileinput
import platform #Windows/Linux
import pymysql #MySQL
import openpyxl #Excel
from openpyxl.styles import Color, PatternFill
import pyminizip #Zips
import shutil #Directory zips
import os
from os import curdir, sep
import calendar
import time
import cgi
import csv
import json
import re
import google_drive_integration
from google.oauth2 import service_account
import random
import email_sender
import ftp_integration

#REMOVE THESE LATER
if (platform.system() == 'Windows'):
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "C:\\Users\\Stefan\\Downloads\\keys\\spannerkey.json" #temp
else:
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "key.json"
spanner_client = spanner.Client()
instance_id = 'marketing-town-reporting'
instance = spanner_client.instance(instance_id)
database_id = 'aditus_crm'
database = instance.database(database_id)

old_dialler_spanner_key = """OLD SPANNER KEY HERE"""

new_dialler_spanner_key = """NEW sSPANNER KEY HERE"""


class fulfilmenttool:

    def get_correct_database_credentials(type):
        #NEW DIALLER
        if "new_dialler" in type:
            service_account_info = json.loads(new_dialler_spanner_key, strict=False)
            credentials = service_account.Credentials.from_service_account_info(service_account_info)
            spanner_client = spanner.Client(project='marketingtown', credentials=credentials)
            instance = spanner_client.instance('marketing-town')
            database = instance.database('marketing-town')
            return database
        #OLD DIALLER
        else:
            service_account_info = json.loads(old_dialler_spanner_key, strict=False)
            credentials = service_account.Credentials.from_service_account_info(service_account_info)
            spanner_client = spanner.Client(project='tech-essence', credentials=credentials)
            instance = spanner_client.instance('marketing-town-reporting')
            database = instance.database('aditus_crm')
            return database

    def do_fulfillment(type,fromdate,todate,datestring,apimode,apiFolder):
        try:
            os.remove("output.zip")
        except:
            fulfilmenttool.update_logs("starting", datestring)
        foldername = datestring
        os.mkdir(foldername)
        config = fulfilmenttool.get_mysql_data(type, "fulfillment")
        print("data is: " + json .dumps(config))
        name, headers, split, pivots, sql_code, pivot_sql_code, header_code, encryption, password, format, misc1, more_than_40_columns, filename, frees_filename, frees_sql_code, frees_format, frees_header_code, ftp_mode, email_recipients = config[0]
        fulfilmenttool.update_logs("Aquired fulfilment configuration options.", datestring)
        filename = fulfilmenttool.substitute_variable(filename)
        frees_filename = fulfilmenttool.substitute_variable(frees_filename)
        sql_code = fulfilmenttool.substitute_code(sql_code, fromdate, todate)
        pivot_sql_code = fulfilmenttool.substitute_code(pivot_sql_code, fromdate, todate)
        frees_sql_code = fulfilmenttool.substitute_code(frees_sql_code, fromdate, todate)
        print ("code is: " + sql_code)
        fulfilmenttool.update_logs("running fulfillment files", datestring)
        NoSalesRows = fulfilmenttool.run_fulfillment_file(sql_code, headers, header_code, filename, format, foldername, more_than_40_columns, datestring, ftp_mode)
        NoFreesRows = fulfilmenttool.split_files(split, frees_filename, frees_sql_code, frees_format, frees_header_code, foldername, more_than_40_columns, datestring, ftp_mode)
        print("NoSalesRows = " + str(NoSalesRows) + ", NoFreesRows = " + str(NoFreesRows))
        if(NoSalesRows == True and NoFreesRows == True):
            fulfilmenttool.update_logs("No files to prepare", datestring)
            fulfilmenttool.tidyup(foldername, apimode, apiFolder, ftp_mode, email_recipients, password, True)
            return
        if (pivots == "Y" or pivots == "Yes"):
            if "RCN" in type:
                fulfilmenttool.build_rcn_pivots(pivot_sql_code, foldername, datestring)
            else:
                fulfilmenttool.build_pivots(pivot_sql_code, foldername, datestring)
        fulfilmenttool.perform_3rd_statement(misc1, fromdate, todate, datestring) 
        fulfilmenttool.encrypt_files(encryption, password, filename + "." + format, split, frees_filename + "." + frees_format, foldername, datestring)
        fulfilmenttool.tidyup(foldername, apimode, apiFolder, ftp_mode, email_recipients, password, False)
        return
  
    def perform_3rd_statement(code, fromdate, todate, datestring):
        pattern = '^update'
        try:
            isPresent = re.match(pattern, code, re.IGNORECASE)
            if isPresent:
                fulfilmenttool.update_logs("running a third statement", datestring)
                adjusted_code = fulfilmenttool.substitute_code(code, fromdate, todate)
                fulfilmenttool.run_spanner_query(adjusted_code, datestring)
                fulfilmenttool.update_logs("third statement run", datestring)
            else:
                fulfilmenttool.update_logs("no third statement", datestring)
        except:
            fulfilmenttool.update_logs("no third statement", datestring)

    def get_mysql_data(type, mode):
        query = ''
        if (mode == "list"):
            query = "SELECT name FROM fulfilment_tool_data"
        else:
            query = "SELECT * FROM fulfilment_tool_data where name = '" + type + "'"
        database = fulfilmenttool.get_correct_database_credentials("")
        with database.snapshot() as snapshot:
             results = snapshot.execute_sql(query)  
        output = []
        for row in results: 
                if len(row) > 1:
                    output.append(row)
                else:
                    outputRow = []
                    outputRow.append(row[0])
                    output.append(outputRow)
        return output


    def update_logs(message, datestring):
        print(message)
        with open("output" + datestring + ".txt", "a") as outputfile:
            outputfile.write("\r\n")
            outputfile.write(message)
        return

    def substitute_code(code, fromdate, todate):
        newcode = code.replace("FROMDATE", fromdate).replace("TODATE", todate)
        return newcode

    def substitute_variable(variable):
        day = datetime.today().strftime('%d')
        month = datetime.today().strftime('%m')
        year = datetime.today().strftime('%y')
        variable2 = variable.replace("IMMEDIATE","JJJJJ") #preventing 'Immediate Media' corruption
        variable3 = variable2.replace("DD",day).replace("MM",month).replace("YY",year)
        variable4 = variable3.replace("JJJJJ","IMMEDIATE")
        return variable4


    def split_files(split, frees_filename, frees_sql_code, frees_format, frees_header_code, foldername, more_than_40_columns, datestring, ftp_mode):
        if (split == 'N'):
            return True
        if (split == 'Y' or split == 'Yes'):
            NoFreesRows = fulfilmenttool.run_fulfillment_file(frees_sql_code, 'Y', frees_header_code, frees_filename, frees_format, foldername, more_than_40_columns, datestring, ftp_mode)
            fulfilmenttool.update_logs("Frees file run.", datestring)
        return NoFreesRows

    def build_rcn_pivots(pivot_sql_code, foldername, datestring): #RCN
        fulfilmenttool.run_fulfillment_file(pivot_sql_code, 'Y', "RCNi,Count", "pivottemp", "csv", foldername, 'N', datestring, None)
        fulfilmenttool.update_logs("running RCN pivots", datestring)
        os.system("echo \"<head><style>table{border-collapse:collapse;} table,th,td {vertical-align: top;border: 1px solid black;text-align: left;}</style></head>\" > " + foldername + "/pivots.html")
        os.system("echo PivotTable >> " + foldername + "/pivots.html")
        filedata = ''
        with open(foldername + '/pivots.html', 'r') as file :
            filedata = file.read()
            filedata = filedata.replace('"', '') #Removing excessive quotes
        file.close()
        with open(foldername + '/pivots.html', 'w') as file:
            file.write(filedata)
        file.close()
        pivotcommand = "pivotdata.exe -i csv " + foldername + "/pivottemp.csv \"{Delimiter:','}\" -p \"{Dimensions:['RCNi', 'Count'],Aggregators:[{Name:'Count'}]}\" -t \"{Rows:['RCNi', 'Count'],Columns:[],AggregatorIndex:1}\" -o pivottablehtml - >> "  + foldername + "/pivots.html"
        if (platform.system() == 'Windows'):
            os.system(pivotcommand)
        else:
            os.system('mono ' + pivotcommand)
        try:
            os.remove(foldername + "/pivottemp.csv")
        except:
            fulfilmenttool.update_logs("No pivots!", datestring)
        fulfilmenttool.update_logs("RCN pivots run", datestring)
        return

    def build_pivots(pivot_sql_code, foldername, datestring):
        fulfilmenttool.run_fulfillment_file(pivot_sql_code, 'Y', "publisher,magazine,fulfillment_code,price", "pivottemp", "csv", foldername, 'N', datestring, None)
        fulfilmenttool.update_logs("running pivots", datestring)
        os.system("echo \"<head><style>table{border-collapse:collapse;} table,th,td {vertical-align: top;border: 1px solid black;text-align: left;}</style></head>\" > " + foldername + "/pivots.html")
        os.system("echo PivotTable >> " + foldername + "/pivots.html")
        filedata = ''
        with open(foldername + '/pivots.html', 'r') as file :
            filedata = file.read()
            filedata = filedata.replace('"', '') #Removing excessive quotes
        file.close()
        with open(foldername + '/pivots.html', 'w') as file:
            file.write(filedata)
        file.close()
        pivotcommand = "pivotdata.exe -i csv " + foldername + "/pivottemp.csv \"{Delimiter:','}\" -p \"{Dimensions:['publisher','magazine', 'fulfillment_code','price'],Aggregators:[{Name:'Count'},{Name:'Sum',Params:['price']}]}\" -t \"{Rows:['publisher','magazine', 'fulfillment_code','price'],Columns:[],AggregatorIndex:1}\" -o pivottablehtml - >> "  + foldername + "/pivots.html"
        if (platform.system() == 'Windows'):
            os.system(pivotcommand)
        else:
            os.system('mono ' + pivotcommand)
        try:
            os.remove(foldername + "/pivottemp.csv")
        except:
            fulfilmenttool.update_logs("No pivots!", datestring)
        fulfilmenttool.update_logs("pivots run", datestring)
        return

    def encrypt_files(encryption, password, file, split, frees_file, foldername, datestring):
        try:
            if (encryption == 'zip'):
                password = fulfilmenttool.substitute_variable(password)
                pyminizip.compress(foldername + "/" + file, "", foldername + "/" + file + ".zip", password, 5)
                if (split == 'Y'):
                    pyminizip.compress(foldername + "/" + frees_file,"",foldername + "/" + frees_file + ".zip", password)
            elif (encryption == 'cds'):
                os.system("gpg --encrypt --recipient thirdparties@cdsglobal.co.uk " + foldername + "/" + file)
                if (split == 'Y'):
                      os.system("gpg --encrypt --recipient thirdparties@cdsglobal.co.uk " + foldername + "/" + frees_file)
            else:
                return
        except:
            fulfilmenttool.update_logs("No files to encrypt.", datestring)
            return
        fulfilmenttool.update_logs("Files encrypted.", datestring)
        return

    def current_publishing_fix(wr, row):
        newRow = []
        i = 1
        for x in row:
            i = i + 1
            if  3 < i < 11:
                newRow.append(x.title())
            else:
                newRow.append(x)
        wr.writerow(newRow)
       
    def investors_chronicle_lookup(folder):
        try:
            file = ''
            for fileName in os.listdir(folder):
                print("Checking: " + fileName)
                if "IC_Link" in fileName:
                    print("found: " + fileName + " to check.")
                    file = fileName
                    break
            if file == '':
                print("No sales file.")
                return
            checkData = ftp_integration.ftp_integration.investors_download(folder)
            workbook = openpyxl.load_workbook(folder + "/" + file)
            worksheet = workbook.active
            rownumber = 0
            matches = 0
            for rows in worksheet.iter_rows(min_row=1, max_row=700, min_col=1):
                rownumber = rownumber + 1
                checkValue = ''
                try:
                    checkValue = str(worksheet.cell(rownumber, 8).value + worksheet.cell(rownumber, 14).value).upper().replace(" ","").replace('"','')
                    print("CHECKVALUE IS: " + checkValue)
                except Exception:
                    print("End of rows")
                    break
                if checkValue in checkData: #Vlookup against active subscribers
                    print("Found a match: " + checkValue)
                    matches = matches + 1
                    for cell in rows:
                      cell.fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type = "solid")
            workbook.save(folder + "/" + file)
            if matches == 0:
                print("No matches.")
        except Exception as e6:
            email_sender.send_error_message("Investors lookup", folder, e6)


    def run_fulfillment_file(sql_code, headers, headercode, filename, format, foldername, more_than_40_columns, datestring, ftp_mode):
        noDataMode = False;
        results = []
        try:
            results = fulfilmenttool.run_spanner_query(sql_code, datestring)
        except Exception as e1:
            email_sender.send_error_message("Spanner", foldername, e1)
        tempfile = foldername + "/tempfile.csv"
        try:
            os.remove(foldername + "/tempfile.csv")
        except Exception as e:
            print("no tempfile")
        index = 0
        for row in results:
                with open(tempfile, 'a', encoding='utf-8', newline='') as myfile:
                    wr = csv.writer(myfile, delimiter=',')
                    if (index == 0 and headers == 'Y'):
                        wr.writerow(headercode.split(","))    
                    if (more_than_40_columns == 'Y'):
                        newRow = row[0].split(',')      
                        wr.writerow(newRow)
                    else:
                        if ("SALES_CURRENT" in filename): #CURRENT
                            fulfilmenttool.current_publishing_fix(wr,row)
                            #wr.writerow([x.upper() for x in row])
                        else:
                            wr.writerow(row)
                    index += 1
        if (ftp_mode == "ABS" and index < 4 and format == 'txt'): #Accounting for when there are only footer rows for ABS.
            print("No rows for ABS")
            noDataMode = True
        if (format == 'txt'):
            for line in fileinput.input([tempfile], inplace=True):
                 print(line.replace('"', ''), end='')
        if (format == 'csv' or format == 'txt'):
            try:
                os.rename(tempfile, foldername + "/" + filename + "." + format)
            except:
                fulfilmenttool.update_logs("No rows for this option!", datestring)
                noDataMode = True
        elif (format == 'xlsx'):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                with open(tempfile, encoding='utf-8') as f:
                    reader = csv.reader(f, delimiter=',')
                    for row in reader:
                        ws.append(row)
                wb.save( foldername + "/" + filename + "." + format)
                os.remove(tempfile)
            except:
                fulfilmenttool.update_logs("No rows for this option!", datestring)
                noDataMode = True
        fulfilmenttool.update_logs("processed file: " + foldername + "/" + filename + "." + format, datestring)
        return noDataMode

    def tidyup(foldername, apimode, apiFolder, ftp_mode, email_recipients, password, noDataMode):
        if noDataMode == True: #Handling Blank Data cases
                for file in os.listdir(foldername):
                    os.remove(foldername + "/" + file)
                f = open(foldername + "/No_data_for_this_option.txt", "x")
                f.close()
        if apimode == True:
            if "nvestors" in apiFolder:
                print("Running Investors checks")
                fulfilmenttool.investors_chronicle_lookup(foldername)
            print("sending to Google Drive")   
            drive_attempts = 0
            while drive_attempts < 4:
                try: #FISH
                    google_drive_integration.google_drive_integration.perform_upload(foldername, apiFolder)
                    break
                except Exception as e3:
                    drive_attempts += 1
                    print("Drive Error number: " + str(drive_attempts))
                    if drive_attempts > 2:
                        email_sender.send_error_message("Google Drive",apiFolder, e3)
                        break
            if noDataMode == False: 
                if ftp_mode is not None and ftp_mode != "": #FTP upload
                    print("Preparing FTP send")
                    if ftp_mode == 'RCN':
                        print("RCN upload!")
                        password = ""
                    ftp_attempts = 0
                    while ftp_attempts < 4:
                        ftp_status = ftp_integration.ftp_integration.prepare_ftp_send(ftp_mode,foldername)
                        if ftp_status == False:
                            ftp_attempts += 1
                            print("FTP Error number: " + str(ftp_attempts))
                            if ftp_attempts >= 3:
                                mail_sender.send_error_message("FTP", foldername, "FTP error")
                                break
                        else:
                            break
                print("Preparing emails")
                if email_recipients is None or email_recipients == '':
                    email_recipients = "stefan@m-t.io,stefan4@m-t.io|stefan5@m-t.io"   #Override
                fixedPassword = ''    
                try:   
                    fixedPassword = fulfilmenttool.substitute_variable(password)
                except Exception as e:
                    True
                email_sender.email_sender.prepare_email(email_recipients, foldername, apiFolder, fixedPassword)
            os.remove("output" + foldername + ".txt")
        else:
            shutil.make_archive("output" + foldername, 'zip', foldername) 
            fulfilmenttool.update_logs("Finished!", foldername)
        shutil.rmtree(foldername)
        print("Finished with " + foldername)
        return

    def a(transaction, sql_code):
        row_ct = transaction.execute_update(sql_code)

    def run_spanner_query(sql_code, datestring):
        pattern = '^update'
        database = fulfilmenttool.get_correct_database_credentials(sql_code)
        isUpdate = re.match(pattern, sql_code, re.IGNORECASE);
        if isUpdate:
            database = fulfilmenttool.get_correct_database_credentials("new_dialler") #Forcing the new dialler to be used for update statements.
            print("Running update statement")
            fulfilmenttool.update_logs("Running UPDATE statement.", datestring)
            database.run_in_transaction(fulfilmenttool.perform_sql_update, sql_code);
            return ''
        else:
            with database.snapshot() as snapshot:
                results = snapshot.execute_sql(sql_code)  
            return results

    def get_titles(request):
        output = fulfilmenttool.get_mysql_data("none","list")
        print("output is: " + json. dumps(output))
        fulfilmenttool.sendResponse(request,"text/html", json. dumps(output))
        return

    #Extracts a file from a HTTP request
    def parse_request(data, apimode):
        try:
            datestring = fulfilmenttool.get_parameter(data, "datestring")
        except:
            if apimode == True:
              datestring = str(round(int(time.time()) * random.uniform(0, 1)))
            else:
               return "Date String missing"
        try:
            os.system("echo '' > output" + datestring + ".txt")
            outcome = "File uploading"
            type = fulfilmenttool.get_parameter(data, "type")
            fromdate = ""
            todate = ""
            apiFolder = ""
            if apimode == True: #Handling an API request where no absolute date range is specified
                timeNow = datetime.now(timezone.utc)
                daysAgo = fulfilmenttool.get_parameter(data, "days")
                startTime = timeNow - timedelta(days=int(daysAgo))     
                fromdate = startTime.strftime('%Y-%m-%d')
                todate=timeNow.strftime('%Y-%m-%d')
                apiFolder = fulfilmenttool.get_parameter(data, "folder") #Output folder on Cloud storage
            else:
                fromdate = fulfilmenttool.get_parameter(data, "fromdate")
                todate = fulfilmenttool.get_parameter(data, "todate")
            fulfilmenttool.multi_process(type,fromdate,todate, datestring, apimode, apiFolder)
        except Exception as e:
            outcome = "Error in processing: " + str(e)
            print("Error: " + str(e))
        return outcome

    def get_parameter(request, parameter):
        parsed = urlparse(request.path)
        return parse_qs(parsed.query)[parameter][0]


    def multi_process(type,fromdate,todate, datestring, apimode, apiFolder):
        proc = Process(target=fulfilmenttool.do_fulfillment, args=(type,fromdate,todate,datestring,apimode,apiFolder)) #Calls the ProcessFile method to run in its own thread
        procs.append(proc)
        proc.start()


    def sendResponse(request,contentType,output):
            request.send_response(200)
            request.send_header('Access-Control-Allow-Origin', '*')
            request.send_header('Content-type',contentType)
            request.end_headers()
            if isinstance(output, str):
                request.wfile.write(output.encode('utf-8'))
            else:
                request.wfile.write(output)
            return

    def download_file(request,filename,contentType):
            try:
                request.send_response(200)
                request.send_header('Access-Control-Allow-Origin', '*')
                request.send_header('Content-type', contentType)
                request.send_header('Content-Disposition', 'attachment; filename="' + filename + '"')
                request.end_headers()
                print("file: " + os.getcwd() + "/" + filename)
                with open(os.getcwd() + "/" + filename, 'rb') as file:
                    request.wfile.write(file.read()) #Download the file
                os.remove(os.getcwd() + "/" + filename)
            except Exception as e:
                    print("Download error: " + str(e));
                    fulfilmenttool.sendResponse(request,"text/html", "Not yet ready.")



      #These two classes handle HTTP requests
class Handler(BaseHTTPRequestHandler):

            def do_GET(self):
                if "gettitles" in self.path:
                    fulfilmenttool.get_titles(self)
                    return
                elif "template" in self.path: #serve a template request
                    fulfilmenttool.download_file(self,"template.csv","text/csv")
                    return
                elif "apirequest" in self.path: #Handles requests from scheduler
                    fulfilmenttool.parse_request(self, True);
                    fulfilmenttool.sendResponse(self,"text/html","File processing")
                    return
                datestring = ""
                try:
                    datestring = fulfilmenttool.get_parameter(self, "datestring")
                except:
                    datestring = ""
                if datestring is None or datestring == "": #Serve the web page 
                    self.path="/index.html" #Page location
                    f = open(curdir + sep + self.path)
                    fulfilmenttool.sendResponse(self,"text/html",f.read())
                    f.close()
                else:
                    try:
                        with open('output' + datestring + '.txt', 'r') as f:
                            lines = f.read().splitlines()
                            outputvalue = lines[-1]
                    except: #No previous file exists
                        fulfilmenttool.sendResponse(self,"text/html","Ready for upload")
                        return
                    if "update" in self.path: #serve an update request
                        if "inished" in outputvalue: #has the process finished?
                            fulfilmenttool.sendResponse(self,"text/html","Please press Download")
                        else:
                            fulfilmenttool.sendResponse(self, "text/html", outputvalue) #Send back the file
                        return
                    elif "download" in self.path: #serve a download request
                        fulfilmenttool.download_file(self,"output" + datestring + ".zip","application/zip")
                        try: #clear completed files
                            os.remove("output" + datestring + ".zip")    
                        except:
                            True
                        try:
                            os.remove("output" + datestring + ".txt")
                        except:
                            True
                    else: 
                        return

            #Handles POST requests from the web page's AJAX calls
            def do_POST(self):
                outputString = fulfilmenttool.parse_request(self, False);
                fulfilmenttool.sendResponse(self,"text/html","File processing")
                return


    #Runs a server
class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
   """Handle requests in a separate thread."""
if __name__ == '__main__':
    procs = []
    server = ThreadedHTTPServer(('0.0.0.0', 8082), Handler) #Serves a server on port 8080.
    print('Starting server, use <Ctrl-C> to stop') #The console message upon running the server.
    #os.system("echo Ready for upload > output.txt"); #Prepares system for next download
    server.serve_forever()
  
